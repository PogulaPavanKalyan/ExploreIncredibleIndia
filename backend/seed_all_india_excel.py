import os
import re
import django
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.states.models import State
from apps.cities.models import City
from apps.categories.models import Category
from apps.destinations.models import Destination
from django.utils.text import slugify

# State Metadata Mapping (Code, Region, Capital)
STATE_METADATA = {
    'Andhra Pradesh': {'code': 'AP', 'region': 'south-india', 'capital': 'Amaravati', 'img': 'https://images.unsplash.com/photo-1627894483216-2138af692e32?w=1200'},
    'Arunachal Pradesh': {'code': 'AR', 'region': 'northeast-india', 'capital': 'Itanagar', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Assam': {'code': 'AS', 'region': 'northeast-india', 'capital': 'Dispur', 'img': 'https://images.unsplash.com/photo-1561731216-c3a4d99437d5?w=1200'},
    'Bihar': {'code': 'BR', 'region': 'east-india', 'capital': 'Patna', 'img': 'https://images.unsplash.com/photo-1590050752117-238cb0fb12b1?w=1200'},
    'Chhattisgarh': {'code': 'CG', 'region': 'central-india', 'capital': 'Raipur', 'img': 'https://images.unsplash.com/photo-1432405972618-c60b0225b8f9?w=1200'},
    'Goa': {'code': 'GA', 'region': 'west-india', 'capital': 'Panaji', 'img': 'https://images.unsplash.com/photo-1512343879784-a960bf40e7f2?w=1200'},
    'Gujarat': {'code': 'GJ', 'region': 'west-india', 'capital': 'Gandhinagar', 'img': 'https://images.unsplash.com/photo-1600100397608-f010e423b971?w=1200'},
    'Haryana': {'code': 'HR', 'region': 'north-india', 'capital': 'Chandigarh', 'img': 'https://images.unsplash.com/photo-1582510003544-4d00b7f74220?w=1200'},
    'Himachal Pradesh': {'code': 'HP', 'region': 'north-india', 'capital': 'Shimla', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Jharkhand': {'code': 'JH', 'region': 'east-india', 'capital': 'Ranchi', 'img': 'https://images.unsplash.com/photo-1432405972618-c60b0225b8f9?w=1200'},
    'Karnataka': {'code': 'KA', 'region': 'south-india', 'capital': 'Bengaluru', 'img': 'https://images.unsplash.com/photo-1600100397608-f010e423b971?w=1200'},
    'Kerala': {'code': 'KL', 'region': 'south-india', 'capital': 'Thiruvananthapuram', 'img': 'https://images.unsplash.com/photo-1602216056096-3b40cc0c9944?w=1200'},
    'Madhya Pradesh': {'code': 'MP', 'region': 'central-india', 'capital': 'Bhopal', 'img': 'https://images.unsplash.com/photo-1589308078059-be1415eab4c3?w=1200'},
    'Maharashtra': {'code': 'MH', 'region': 'west-india', 'capital': 'Mumbai', 'img': 'https://images.unsplash.com/photo-1570168007204-dfb528c6958f?w=1200'},
    'Manipur': {'code': 'MN', 'region': 'northeast-india', 'capital': 'Imphal', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Meghalaya': {'code': 'ML', 'region': 'northeast-india', 'capital': 'Shillong', 'img': 'https://images.unsplash.com/photo-1432405972618-c60b0225b8f9?w=1200'},
    'Mizoram': {'code': 'MZ', 'region': 'northeast-india', 'capital': 'Aizawl', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Nagaland': {'code': 'NL', 'region': 'northeast-india', 'capital': 'Kohima', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Odisha': {'code': 'OR', 'region': 'east-india', 'capital': 'Bhubaneswar', 'img': 'https://images.unsplash.com/photo-1600100397608-f010e423b971?w=1200'},
    'Punjab': {'code': 'PB', 'region': 'north-india', 'capital': 'Chandigarh', 'img': 'https://images.unsplash.com/photo-1588096344356-9a2503a74313?w=1200'},
    'Rajasthan': {'code': 'RJ', 'region': 'north-india', 'capital': 'Jaipur', 'img': 'https://images.unsplash.com/photo-1477587458883-47145ed94245?w=1200'},
    'Sikkim': {'code': 'SK', 'region': 'northeast-india', 'capital': 'Gangtok', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Tamil Nadu': {'code': 'TN', 'region': 'south-india', 'capital': 'Chennai', 'img': 'https://images.unsplash.com/photo-1582510003544-4d00b7f74220?w=1200'},
    'Telangana': {'code': 'TS', 'region': 'south-india', 'capital': 'Hyderabad', 'img': 'https://images.unsplash.com/photo-1605649487212-47bdab064df7?w=1200'},
    'Tripura': {'code': 'TR', 'region': 'northeast-india', 'capital': 'Agartala', 'img': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200'},
    'Uttar Pradesh': {'code': 'UP', 'region': 'north-india', 'capital': 'Lucknow', 'img': 'https://images.unsplash.com/photo-1564507592333-c60657eea523?w=1200'},
    'Uttarakhand': {'code': 'UK', 'region': 'north-india', 'capital': 'Dehradun', 'img': 'https://images.unsplash.com/photo-1626621341517-bbf3d9990a23?w=1200'},
    'West Bengal': {'code': 'WB', 'region': 'east-india', 'capital': 'Kolkata', 'img': 'https://images.unsplash.com/photo-1558431382-27e303142255?w=1200'}
}

# District Resolution Helper Dictionary for Known Famous Places across India
DISTRICT_LOOKUP = {
    # Andhra Pradesh
    'Tirupati': 'Tirupati', 'Sri Venkateswara Temple': 'Tirupati', 'Srikalahasti Temple': 'Tirupati',
    'Kanaka Durga Temple': 'NTR / Vijayawada', 'Araku Valley': 'Alluri Sitharama Raju',
    'Borra Caves': 'Alluri Sitharama Raju', 'Simhachalam Temple': 'Visakhapatnam', 'Rk Beach': 'Visakhapatnam',
    'Submarine Museum': 'Visakhapatnam', 'Kailasagiri': 'Visakhapatnam', 'Lepakshi': 'Sri Sathya Sai',
    'Belum Caves': 'Nandyal', 'Yaganti': 'Nandyal', 'Srisailam Temple': 'Nandyal',
    'Horsley Hills': 'Annamayya', 'Gandikota': 'YSR Kadapa', 'Mypadu Beach': 'Nellore',
    'Suryalanka Beach': 'Bapatla', 'Bhavani Island': 'NTR / Vijayawada', 'Kondapalli Fort': 'NTR / Vijayawada',

    # Himachal Pradesh
    'Shimla': 'Shimla', 'Kufri': 'Shimla', 'Jakhoo Temple': 'Shimla', 'Mall Road Shimla': 'Shimla',
    'Manali': 'Kullu', 'Solang Valley': 'Kullu', 'Rohtang Pass': 'Kullu', 'Atal Tunnel': 'Kullu',
    'Kullu': 'Kullu', 'Kasol': 'Kullu', 'Manikaran': 'Kullu', 'Dharamshala': 'Kangra',
    'McLeod Ganj': 'Kangra', 'Bhagsunag Waterfall': 'Kangra', 'Cricket Stadium Dharamshala': 'Kangra',
    'Dalhousie': 'Chamba', 'Khajjiar': 'Chamba', 'Spiti Valley': 'Lahaul and Spiti', 'Kaza': 'Lahaul and Spiti',
    'Key Monastery': 'Lahaul and Spiti', 'Chandratal Lake': 'Lahaul and Spiti', 'Tabo Monastery': 'Lahaul and Spiti',
    'Pin Valley National Park': 'Lahaul and Spiti', 'Sangla Valley': 'Kinnaur', 'Kalpa': 'Kinnaur',

    # Kerala
    'Munnar': 'Idukki', 'Tea Gardens Munnar': 'Idukki', 'Anamudi Peak': 'Idukki', 'Mattupetty Dam': 'Idukki',
    'Alleppey': 'Alappuzha', 'Alleppey Backwaters': 'Alappuzha', 'Vembanad Lake': 'Alappuzha',
    'Kochi': 'Ernakulam', 'Fort Kochi': 'Ernakulam', 'Mattancherry Palace': 'Ernakulam',
    'Wayanad': 'Wayanad', 'Edakkal Caves': 'Wayanad', 'Banasura Sagar Dam': 'Wayanad', 'Chembra Peak': 'Wayanad',
    'Varkala': 'Thiruvananthapuram', 'Varkala Beach': 'Thiruvananthapuram', 'Kovalam Beach': 'Thiruvananthapuram',
    'Padmanabhaswamy Temple': 'Thiruvananthapuram', 'Thekkady': 'Idukki', 'Periyar National Park': 'Idukki',
    'Athirappilly Waterfalls': 'Thrissur', 'Kumarakom': 'Kottayam', 'Bekal Fort': 'Kasaragod',

    # Rajasthan
    'Jaipur': 'Jaipur', 'Amer Fort': 'Jaipur', 'Hawa Mahal': 'Jaipur', 'City Palace Jaipur': 'Jaipur',
    'Jantar Mantar Jaipur': 'Jaipur', 'Nahargarh Fort': 'Jaipur', 'Jodhpur': 'Jodhpur',
    'Mehrangarh Fort': 'Jodhpur', 'Umaid Bhawan Palace': 'Jodhpur', 'Jaswant Thada': 'Jodhpur',
    'Udaipur': 'Udaipur', 'City Palace Udaipur': 'Udaipur', 'Lake Pichola': 'Udaipur', 'Jag Mandir': 'Udaipur',
    'Jaisalmer': 'Jaisalmer', 'Jaisalmer Fort': 'Jaisalmer', 'Sam Sand Dunes': 'Jaisalmer',
    'Pushkar': 'Ajmer', 'Pushkar Lake': 'Ajmer', 'Brahma Temple Pushkar': 'Ajmer',
    'Ranthambore National Park': 'Sawai Madhopur', 'Mount Abu': 'Sirohi', 'Dilwara Temples': 'Sirohi',
    'Chittorgarh Fort': 'Chittorgarh', 'Bikaner': 'Bikaner', 'Junagarh Fort': 'Bikaner',

    # Uttar Pradesh
    'Taj Mahal': 'Agra', 'Agra Fort': 'Agra', 'Fatehpur Sikri': 'Agra', 'Varanasi': 'Varanasi',
    'Kashi Vishwanath Temple': 'Varanasi', 'Dashashwamedh Ghat': 'Varanasi', 'Sarnath': 'Varanasi',
    'Mathura': 'Mathura', 'Vrindavan': 'Mathura', 'Banke Bihari Temple': 'Mathura', 'Prem Mandir': 'Mathura',
    'Ayodhya': 'Ayodhya', 'Ram Mandir Ayodhya': 'Ayodhya', 'Hanuman Garhi': 'Ayodhya',
    'Lucknow': 'Lucknow', 'Bada Imambara': 'Lucknow', 'Chota Imambara': 'Lucknow', 'Prayagraj': 'Prayagraj',
    'Triveni Sangam Prayagraj': 'Prayagraj', 'Kushinagar': 'Kushinagar', 'Jhansi Fort': 'Jhansi',

    # Uttarakhand
    'Kedarnath Temple': 'Rudraprayag', 'Badrinath Temple': 'Chamoli', 'Gangotri Temple': 'Uttarkashi',
    'Yamunotri Temple': 'Uttarkashi', 'Rishikesh': 'Dehradun', 'Laxman Jhula': 'Dehradun', 'Triveni Ghat Rishikesh': 'Dehradun',
    'Haridwar': 'Haridwar', 'Har Ki Pauri': 'Haridwar', 'Nainital': 'Nainital', 'Naini Lake': 'Nainital',
    'Mussoorie': 'Dehradun', 'Kempty Falls': 'Dehradun', 'Auli': 'Chamoli', 'Jim Corbett National Park': 'Nainital',
    'Valley Of Flowers': 'Chamoli', 'Ranikhet': 'Almora', 'Kausani': 'Bageshwar', 'Mukteshwar': 'Nainital',

    # Karnataka
    'Bengaluru': 'Bengaluru Urban', 'Lalbagh Botanical Garden': 'Bengaluru Urban', 'Cubbon Park': 'Bengaluru Urban',
    'Bangalore Palace': 'Bengaluru Urban', 'Mysuru': 'Mysuru', 'Mysore Palace': 'Mysuru', 'Chamundi Hill': 'Mysuru',
    'Hampi': 'Vijayanagara', 'Virupaksha Temple': 'Vijayanagara', 'Stone Chariot Hampi': 'Vijayanagara',
    'Coorg': 'Kodagu', 'Abbey Falls': 'Kodagu', 'Raja Seat Coorg': 'Kodagu', 'Gokarna': 'Uttara Kannada',
    'Om Beach': 'Uttara Kannada', 'Chikmagalur': 'Chikkamagaluru', 'Mullayanagiri': 'Chikkamagaluru',
    'Badami': 'Bagalkot', 'Pattadakal': 'Bagalkot', 'Aihole': 'Bagalkot', 'Jog Falls': 'Shivamogga',
    'Bandipur National Park': 'Chamarajanagar', 'Udupi': 'Udupi', 'Murudeshwar': 'Uttara Kannada',

    # Maharashtra
    'Mumbai': 'Mumbai City', 'Gateway Of India': 'Mumbai City', 'Marine Drive': 'Mumbai City', 'Elephanta Caves': 'Mumbai Suburban',
    'Pune': 'Pune', 'Shaniwar Wada': 'Pune', 'Aga Khan Palace': 'Pune', 'Lonavala': 'Pune', 'Khandala': 'Pune',
    'Mahabaleshwar': 'Satara', 'Panchgani': 'Satara', 'Ajanta Caves': 'Chhatrapati Sambhaji Nagar',
    'Ellora Caves': 'Chhatrapati Sambhaji Nagar', 'Shirdi': 'Ahilya Nagar / Ahmednagar', 'Sai Baba Temple Shirdi': 'Ahilya Nagar',
    'Nashik': 'Nashik', 'Trimbakeshwar Temple': 'Nashik', 'Tadoba National Park': 'Chandrapur',
    'Alibaug': 'Raigad', 'Ganpatipule': 'Ratnagiri', 'Matheran': 'Raigad',

    # Tamil Nadu
    'Chennai': 'Chennai', 'Marina Beach': 'Chennai', 'Kapaleeshwarar Temple': 'Chennai', 'Mahabalipuram': 'Chengalpattu',
    'Shore Temple': 'Chengalpattu', 'Madurai': 'Madurai', 'Meenakshi Amman Temple': 'Madurai',
    'Rameswaram': 'Ramanathapuram', 'Ramanathaswamy Temple': 'Ramanathapuram', 'Kanyakumari': 'Kanyakumari',
    'Vivekananda Rock Memorial': 'Kanyakumari', 'Ooty': 'Nilgiris', 'Botanical Gardens Ooty': 'Nilgiris',
    'Doddabetta Peak': 'Nilgiris', 'Kodaikanal': 'Dindigul', 'Kodaikanal Lake': 'Dindigul',
    'Thanjavur': 'Thanjavur', 'Brihadeeswarar Temple': 'Thanjavur', 'Kanchipuram': 'Kanchipuram',
    'Chidambaram': 'Cuddalore', 'Coimbatore': 'Coimbatore', 'Adiyogi Shiva Statue': 'Coimbatore',

    # West Bengal
    'Kolkata': 'Kolkata', 'Victoria Memorial': 'Kolkata', 'Howrah Bridge': 'Kolkata', 'Dakshineswar Kali Temple': 'Kolkata',
    'Darjeeling': 'Darjeeling', 'Tiger Hill Darjeeling': 'Darjeeling', 'Darjeeling Himalayan Railway': 'Darjeeling',
    'Kalimpong': 'Kalimpong', 'Sundarbans': 'South 24 Parganas', 'Digha': 'Purba Medinipur', 'Shantiniketan': 'Birbhum',
    'Murshidabad': 'Murshidabad', 'Hazarduari Palace': 'Murshidabad', 'Bishnupur': 'Bankura',

    # Goa
    'Calangute Beach': 'North Goa', 'Baga Beach': 'North Goa', 'Anjuna Beach': 'North Goa', 'Aguada Fort': 'North Goa',
    'Basilica Of Bom Jesus': 'North Goa', 'Dudhsagar Waterfalls': 'South Goa', 'Colva Beach': 'South Goa',
    'Palolem Beach': 'South Goa', 'Panaji': 'North Goa', 'Fontainhas': 'North Goa'
}

# Image Map for Common Places (High-res Unsplash direct URLs)
DEFAULT_CATEGORY_IMAGES = {
    'Religious': 'https://images.unsplash.com/photo-1600100397608-f010e423b971?w=1200',
    'Temples & Spiritual': 'https://images.unsplash.com/photo-1600100397608-f010e423b971?w=1200',
    'Historical & Heritage': 'https://images.unsplash.com/photo-1590050752117-238cb0fb12b1?w=1200',
    'Heritage & Culture': 'https://images.unsplash.com/photo-1590050752117-238cb0fb12b1?w=1200',
    'Hills & Mountains': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200',
    'Hills & Nature': 'https://images.unsplash.com/photo-1506744038136-46273834b3fb?w=1200',
    'Wildlife & Nature': 'https://images.unsplash.com/photo-1561731216-c3a4d99437d5?w=1200',
    'Waterfall': 'https://images.unsplash.com/photo-1432405972618-c60b0225b8f9?w=1200',
    'Waterfalls & Lakes': 'https://images.unsplash.com/photo-1432405972618-c60b0225b8f9?w=1200',
    'Beach': 'https://images.unsplash.com/photo-1512343879784-a960bf40e7f2?w=1200',
    'Beaches & Coastal': 'https://images.unsplash.com/photo-1512343879784-a960bf40e7f2?w=1200',
    'Tourist Attraction': 'https://images.unsplash.com/photo-1582510003544-4d00b7f74220?w=1200'
}

def normalize_category(cat_raw, place_name):
    cat_str = (cat_raw or '').strip()
    name_str = (place_name or '').strip().lower()

    if 'temple' in name_str or 'shrine' in name_str or 'monastery' in name_str or 'gurudwara' in name_str or 'church' in name_str or 'cathedral' in name_str or cat_str == 'Religious':
        return 'Temples & Spiritual'
    elif 'fort' in name_str or 'palace' in name_str or 'mahal' in name_str:
        return 'Forts & Palaces'
    elif 'falls' in name_str or 'waterfall' in name_str or 'lake' in name_str or 'dam' in name_str or cat_str in ['Waterfall', 'Nature & Water']:
        return 'Waterfalls & Lakes'
    elif 'beach' in name_str or 'coast' in name_str or cat_str == 'Beach':
        return 'Beaches & Coastal'
    elif 'park' in name_str or 'sanctuary' in name_str or 'wildlife' in name_str or 'tiger' in name_str or cat_str == 'Wildlife & Nature':
        return 'Wildlife & Nature'
    elif 'valley' in name_str or 'peak' in name_str or 'hill' in name_str or cat_str == 'Hills & Mountains':
        return 'Hills & Nature'
    elif cat_str in ['Historical & Heritage', 'Heritage']:
        return 'Heritage & Culture'
    else:
        return 'Heritage & Culture'

def get_approx_coords(place_name, state_name):
    # Geocoding lookup table
    coords = {
        'Tirupati': (13.6288, 79.4192), 'Sri Venkateswara Temple': (13.6833, 79.3472),
        'Srikalahasti Temple': (13.7498, 79.6984), 'Kanaka Durga Temple': (16.5165, 80.6074),
        'Araku Valley': (18.3273, 82.8775), 'Borra Caves': (18.2804, 83.0384),
        'Shimla': (31.1048, 77.1734), 'Manali': (32.2432, 77.1892), 'Dharamshala': (32.2190, 76.3234),
        'McLeod Ganj': (32.2426, 76.3213), 'Dalhousie': (32.5387, 75.9710), 'Spiti Valley': (32.2461, 78.0349),
        'Munnar': (10.0889, 77.0595), 'Alleppey': (9.4981, 76.3388), 'Kochi': (9.9312, 76.2673),
        'Wayanad': (11.6854, 76.1320), 'Varkala': (8.7379, 76.7163), 'Jaipur': (26.9124, 75.7873),
        'Amer Fort': (26.9855, 75.8513), 'Hawa Mahal': (26.9239, 75.8267), 'Jodhpur': (26.2389, 73.0243),
        'Mehrangarh Fort': (26.2978, 73.0185), 'Udaipur': (24.5854, 73.7125), 'Jaisalmer': (26.9157, 70.9083),
        'Taj Mahal': (27.1751, 78.0421), 'Varanasi': (25.3176, 82.9739), 'Kedarnath Temple': (30.7346, 79.0669),
        'Badrinath Temple': (30.7433, 79.4938), 'Rishikesh': (30.0869, 78.2676), 'Haridwar': (29.9457, 78.1642),
        'Nainital': (29.3919, 79.4542), 'Hampi': (15.3350, 76.4600), 'Coorg': (12.4244, 75.7382),
        'Gokarna': (14.5479, 74.3188), 'Mumbai': (19.0760, 72.8777), 'Ajanta Caves': (20.5519, 75.7033),
        'Ellora Caves': (20.0268, 75.1771), 'Chennai': (13.0827, 80.2707), 'Madurai': (9.9252, 78.1198),
        'Rameswaram': (9.2876, 79.3129), 'Kanyakumari': (8.0883, 77.5385), 'Ooty': (11.4102, 76.6950),
        'Kolkata': (22.5726, 88.3639), 'Darjeeling': (27.0410, 88.2663), 'Panaji': (15.4909, 73.8278)
    }

    if place_name in coords:
        return coords[place_name]
    
    # State base centroids
    state_centroids = {
        'Andhra Pradesh': (15.9129, 79.7400), 'Arunachal Pradesh': (28.2180, 94.7278),
        'Assam': (26.2006, 92.9376), 'Bihar': (25.0961, 85.3131), 'Chhattisgarh': (21.2787, 81.8661),
        'Goa': (15.2993, 74.1240), 'Gujarat': (22.2587, 71.1924), 'Haryana': (29.0588, 76.0856),
        'Himachal Pradesh': (31.1048, 77.1734), 'Jharkhand': (23.6102, 85.2799),
        'Karnataka': (15.3173, 75.7139), 'Kerala': (10.8505, 76.2711), 'Madhya Pradesh': (22.9734, 78.6569),
        'Maharashtra': (19.7515, 75.7139), 'Manipur': (24.6637, 93.9063), 'Meghalaya': (25.4670, 91.3662),
        'Mizoram': (23.1645, 92.9376), 'Nagaland': (26.1584, 94.5624), 'Odisha': (20.9517, 85.0985),
        'Punjab': (31.1471, 75.3412), 'Rajasthan': (27.0238, 74.2179), 'Sikkim': (27.5330, 88.5122),
        'Tamil Nadu': (11.1271, 78.6569), 'Telangana': (18.1124, 79.0193), 'Tripura': (23.9408, 91.9882),
        'Uttar Pradesh': (26.8467, 80.9462), 'Uttarakhand': (30.0668, 79.0193), 'West Bengal': (22.9868, 87.8550)
    }

    base = state_centroids.get(state_name, (20.5937, 78.9629))
    # Slight deterministic scatter based on name length to avoid exact coordinate overlap
    offset_lat = (len(place_name) % 10 - 5) * 0.08
    offset_lng = (len(place_name) % 7 - 3) * 0.08
    return (round(base[0] + offset_lat, 4), round(base[1] + offset_lng, 4))

def seed_excel_dataset():
    print("==================================================")
    print("Starting All India 625 Tourist Places Database Seeding...")
    print("==================================================")

    file_path = '../All_India_Medium_to_Big_Tourist_Places_One_Sheet.xlsx'
    if not os.path.exists(file_path):
        file_path = 'All_India_Medium_to_Big_Tourist_Places_One_Sheet.xlsx'

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    rows = []
    for r in range(2, sheet.max_row + 1):
        row_dict = {headers[i]: sheet.cell(r, i + 1).value for i in range(len(headers))}
        rows.append(row_dict)

    print(f"Loaded {len(rows)} tourist place records from Excel sheet.")

    # 1. Create/Ensure State Objects
    states_created = 0
    state_objects = {}
    for st_name, meta in STATE_METADATA.items():
        st_slug = slugify(st_name)
        st_obj, created = State.objects.get_or_create(
            slug=st_slug,
            defaults={
                'name': st_name,
                'code': meta['code'],
                'capital': meta['capital'],
                'region': meta['region'],
                'image': meta['img'],
                'banner_image': meta['img'],
                'description': f"{st_name} is one of India's premier travel destinations featuring magnificent cultural heritage, sacred temples, natural landscapes, and rich regional traditions.",
                'is_active': True
            }
        )
        state_objects[st_name] = st_obj
        if created:
            states_created += 1

    print(f"Processed {len(state_objects)} States (New: {states_created}).")

    # 2. Seed Categories
    category_names = [
        'Temples & Spiritual', 'Heritage & Culture', 'Forts & Palaces',
        'Hills & Nature', 'Wildlife & Nature', 'Waterfalls & Lakes',
        'Beaches & Coastal', 'Adventure & Trekking', 'Urban & Leisure'
    ]
    category_objects = {}
    for cat_n in category_names:
        c_obj, _ = Category.objects.get_or_create(
            slug=slugify(cat_n),
            defaults={'name': cat_n, 'description': f"Top places for {cat_n}"}
        )
        category_objects[cat_n] = c_obj

    # 3. Seed Destinations & Cities/Districts
    dest_created = 0
    dest_updated = 0

    for item in rows:
        st_name = item.get('State')
        place_name = item.get('Place Name')
        raw_cat = item.get('Category')
        short_desc = item.get('Short Description') or f"Famous tourist destination in {st_name}."
        why_visit = item.get('Why Visit') or ''
        highlights = item.get('Highlights') or ''
        things_to_do = item.get('Things To Do') or ''
        best_time = item.get('Best Time To Visit') or 'October to March'

        if not st_name or not place_name:
            continue

        state_obj = state_objects.get(st_name)
        if not state_obj:
            st_slug = slugify(st_name)
            state_obj, _ = State.objects.get_or_create(
                slug=st_slug,
                defaults={'name': st_name, 'region': 'south-india', 'is_active': True}
            )
            state_objects[st_name] = state_obj

        # Resolve Revenue District
        district_name = DISTRICT_LOOKUP.get(place_name)
        if not district_name:
            # Check if place name has a location hint or fallback to Place Name / Capital
            district_name = place_name.split()[0] if len(place_name.split()) > 1 else place_name

        # Get/Create District City Object
        dist_slug = slugify(f"{district_name}-{st_name}")
        city_obj = City.objects.filter(slug=dist_slug).first()
        if not city_obj:
            city_obj = City.objects.filter(name=district_name, state=state_obj).first()
        if not city_obj:
            try:
                city_obj = City.objects.create(
                    name=district_name,
                    state=state_obj,
                    slug=dist_slug,
                    published=True
                )
            except Exception:
                city_obj = City.objects.filter(state=state_obj).first()

        # Normalize Category
        norm_cat_name = normalize_category(raw_cat, place_name)
        cat_obj = category_objects.get(norm_cat_name)

        # Geocode
        lat, lng = get_approx_coords(place_name, st_name)

        # Cover Image
        img_url = DEFAULT_CATEGORY_IMAGES.get(raw_cat) or DEFAULT_CATEGORY_IMAGES.get(norm_cat_name) or 'https://images.unsplash.com/photo-1582510003544-4d00b7f74220?w=1200'

        # Create or Update Destination
        dest_slug = slugify(f"{place_name}-{st_name}")
        dest_obj, created = Destination.objects.get_or_create(
            slug=dest_slug,
            defaults={
                'name': place_name,
                'state': state_obj,
                'district': district_name,
                'city': city_obj,
                'region': state_obj.region,
                'famous_for': why_visit or short_desc,
                'short_description': short_desc,
                'description': f"{place_name} in {district_name}, {st_name}. {why_visit} {highlights}",
                'things_to_do': things_to_do,
                'best_time_to_visit': best_time,
                'latitude': lat,
                'longitude': lng,
                'main_image': img_url,
                'ticket_price': 0.00,
                'avg_rating': 4.8,
                'published': True
            }
        )

        if cat_obj and not dest_obj.categories.filter(id=cat_obj.id).exists():
            dest_obj.categories.add(cat_obj)

        if created:
            dest_created += 1
        else:
            # Ensure district, city, lat, lng, image are updated
            dest_obj.district = district_name
            dest_obj.city = city_obj
            dest_obj.latitude = lat
            dest_obj.longitude = lng
            if not dest_obj.main_image:
                dest_obj.main_image = img_url
            dest_obj.save()
            dest_updated += 1

    print("\n==================================================")
    print("ALL INDIA DATABASE SEEDING COMPLETED SUCCESSFULLY!")
    print(f"Total States Active: {State.objects.count()}")
    print(f"Total Districts Active: {City.objects.count()}")
    print(f"Total Destinations Seeded: {Destination.objects.count()} (New: {dest_created}, Refined: {dest_updated})")
    print("==================================================")

if __name__ == '__main__':
    seed_excel_dataset()
